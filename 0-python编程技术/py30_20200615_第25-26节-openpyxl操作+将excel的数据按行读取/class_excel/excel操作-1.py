"""
======================
Author: 柠檬班-小简
Time: 2020/6/15 20:29
Project: py30
Company: 湖南零檬信息技术有限公司
======================
"""
"""
excel 

openpyxl: .xlsx读写操作
安装：pip install openpyxl

测试数据：事先会写好在excel
平常操作excel的流程(3个对象)：
    工作薄(Workbook)
    表单(Sheet)
    单元格(Cell)
    
打开一个excel,选择一个表单，在表单里面读取单元格的值。

1、准备测试数据
2、load_workbook模块，去打开测试数据文件，生成WorkBook对象(wb)
3、根据表单名称选择表单(sh)：wb['表单名称']
4、在表单当中，获取单元格的数据：
   4.1 单元格对象：sh.cell(row,colum)  # 下标从1开始
   4.2 .value获取单元格的值。
   4.3 修改数据：sh.cell(row,colum).value = 新的值   

5、得到当前表单当中，总行数和总列数
   sh.max_row # 总行数
   sh.max_column  # 总列数

6、在表单当中，获取单元格的数据：
   6.3 修改数据：sh.cell(row,colum).value = 新的值

7、保存数据(保存整个工作薄)
    WorkBook对象(wb).save(文件路径)
    保存到原文件的时候，需要注意：文件没有被占用，否则会被权限不允许的错误。

所有读取出来的数据：字符串、数字 
"""
import os
file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),"login_cases.xlsx")
# print(file_dir)

# 1、加载excel数据文件
from openpyxl import load_workbook
wb = load_workbook(file_path)

# 2、根据表单名称选择表单：wb['表单名称']
sh = wb["login"]
print(sh)

# 3、单元格对象：sh.cell(row,colum)  # 下标从1开始
cel = sh.cell(2,2)
print(cel.value)

# sh.max_row # 总行数
# sh.max_column  # 总列数
# print(sh.max_row)
# print(sh.max_column)

# print("8888888888888888888888888888888888888888888888888888")
# # 6.3 修改数据：sh.cell(row,colum).value = 新的值
# sh.cell(2,2).value = "lemonban66666"
# print(sh.cell(2,2).value)

# 7、保存数据(保存整个工作薄)
#     WorkBook对象(wb).save(文件路径)
# wb.save("save_as_another_excel.xlsx")  另存在

# wb.save(file_path)